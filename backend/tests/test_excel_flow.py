from pathlib import Path
from io import BytesIO

from openpyxl import load_workbook
from openpyxl import Workbook

from app.schemas.calculation import CalculateRequest, ProjectPayload
from app.schemas.mappers import request_to_engine, result_to_response
from app.services.excel import export_result, import_bids_from_xlsx
from pricing_engine.registry import calculate


ROOT = Path(__file__).resolve().parents[2]


def test_import_calculate_export_with_test1_xlsx() -> None:
    imported = import_bids_from_xlsx((ROOT / "测试1.xlsx").read_bytes())
    assert len(imported.rows) == 25
    assert imported.mapping == {"bidder_name": "投标人名称", "bid_price": "投标价格"}

    payload = CalculateRequest(
        project=ProjectPayload(
            tender_no="T001",
            section_no="S001",
            package_no="P001",
            province="浙江",
            ceiling_price="200",
            price_weight="0.3",
            target_company="投标人23",
        ),
        method_code="A01",
        params={},
        bids=imported.rows,
        source="xlsx_upload",
    )
    response = result_to_response(calculate(request_to_engine(payload))).model_dump()
    content = export_result(payload, response)
    wb = load_workbook(BytesIO(content), read_only=True)
    assert wb.sheetnames == ["测算结果", "报价明细", "参数设置", "计算过程"]
    detail = wb["报价明细"]
    assert detail.max_row == 26
    assert [cell.value for cell in detail[1]] == [
        "投标人名称",
        "投标价格",
        "价格分",
        "排名",
        "是否参与计算",
        "是否用于基准价",
        "备注",
    ]
    ranks = [row[3] for row in detail.iter_rows(min_row=2, values_only=True) if row[3] is not None]
    assert ranks == sorted(ranks)


def test_import_returns_manual_mapping_preview_for_unknown_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公司", "金额"])
    sheet.append(["甲公司", "100"])
    sheet.append(["乙公司", "不开标"])
    stream = BytesIO()
    workbook.save(stream)

    imported = import_bids_from_xlsx(stream.getvalue())

    assert imported.requires_mapping is True
    assert imported.columns == ["公司", "金额"]
    assert imported.preview == [{"公司": "甲公司", "金额": "100"}, {"公司": "乙公司", "金额": "不开标"}]
