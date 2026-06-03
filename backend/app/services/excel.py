from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.calculation import BidPayload, CalculateRequest, ImportResponse

BIDDER_HEADERS = {"投标人名称", "投标人"}
PRICE_HEADERS = {"投标价格", "投标价", "报价"}
MAX_PREVIEW_ROWS = 30


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def import_bids_from_xlsx(content: bytes) -> ImportResponse:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    warnings: list[str] = []
    first_tabular_preview: ImportResponse | None = None
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for row_index, row in enumerate(rows[:30]):
            headers = [_cell_text(value) for value in row]
            non_empty_headers = [header for header in headers if header]
            if len(non_empty_headers) >= 2 and first_tabular_preview is None:
                preview = []
                for data_row in rows[row_index + 1 : row_index + 1 + MAX_PREVIEW_ROWS]:
                    item = {
                        header: _cell_text(data_row[i] if i < len(data_row) else None)
                        for i, header in enumerate(headers)
                        if header
                    }
                    if any(item.values()):
                        preview.append(item)
                if preview:
                    first_tabular_preview = ImportResponse(
                        rows=[],
                        mapping={},
                        warnings=["未自动识别投标人列或报价列，请手动选择列映射"],
                        requires_mapping=True,
                        columns=non_empty_headers,
                        preview=preview,
                    )
            bidder_col = next((i for i, header in enumerate(headers) if header in BIDDER_HEADERS), None)
            price_col = next((i for i, header in enumerate(headers) if header in PRICE_HEADERS), None)
            if bidder_col is None or price_col is None:
                continue
            bid_rows: list[BidPayload] = []
            for data_row in rows[row_index + 1 :]:
                bidder = _cell_text(data_row[bidder_col] if bidder_col < len(data_row) else None)
                price = _cell_text(data_row[price_col] if price_col < len(data_row) else None)
                if not bidder and not price:
                    continue
                if bidder:
                    bid_rows.append(BidPayload(bidder_name=bidder, bid_price=price))
            return ImportResponse(
                rows=bid_rows,
                mapping={"bidder_name": headers[bidder_col], "bid_price": headers[price_col]},
                warnings=warnings,
            )
    if first_tabular_preview is not None:
        return first_tabular_preview
    raise ValueError("未识别投标人列或报价列，请使用标准表头或手动录入")


def _append_pairs(ws: Worksheet, rows: list[tuple[str, Any]]) -> None:
    for key, value in rows:
        ws.append([key, value])


def _ranked_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("rank") is None,
            row.get("rank") if row.get("rank") is not None else 0,
        ),
    )


def export_result(payload: CalculateRequest, result: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "测算结果"
    _append_pairs(
        ws,
        [
            ("招标编号", payload.project.tender_no),
            ("招标名称", payload.project.tender_name or ""),
            ("分标编号", payload.project.section_no),
            ("包号", payload.project.package_no),
            ("招标省份", payload.project.province),
            ("评分方法", f"{result['method_code']} {result['method_name']}"),
            ("基准价", result["benchmark_price"]),
            ("基准价折扣率", result.get("discount_rate") or ""),
            ("参与计算人数", result["bidder_count"]),
            ("有效报价人数", result["effective_count"]),
            ("目标公司排名", result["target"].get("rank") or ""),
            ("目标公司得分", result["target"].get("score") or ""),
            ("与第一名分差", result["target"].get("score_gap") or ""),
            ("折算后分差", result["target"].get("weighted_gap") or ""),
        ],
    )

    detail = wb.create_sheet("报价明细")
    detail.append(["投标人名称", "投标价格", "价格分", "排名", "是否参与计算", "是否用于基准价", "备注"])
    for row in _ranked_rows(result["rows"]):
        detail.append(
            [
                row["bidder_name"],
                row.get("bid_price") or "",
                row.get("score") or "",
                row.get("rank") or "",
                "是" if row["participated"] else "否",
                "是" if row["used_for_benchmark"] else "否",
                row.get("remark") or "",
            ]
        )

    params = wb.create_sheet("参数设置")
    params.append(["参数", "值"])
    for key, value in sorted(payload.params.items()):
        params.append([key, str(value)])

    trace = wb.create_sheet("计算过程")
    trace.append(["字段", "值"])
    for key, value in result["debug"].items():
        trace.append([key, str(value)])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 36)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
