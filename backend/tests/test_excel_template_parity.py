from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pricing_engine.models import BidInput, CalculationInput, ProjectInput
from pricing_engine.registry import calculate

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "国网报价算分模板全8套v1.2.xlsx"
BENCHMARK_TOLERANCE = Decimal("0.000001")
SCORE_TOLERANCE = Decimal("0.0001")

SHEET_METHODS = {
    "区间平均价浮动法": "A01",
    "区间复合平均价法（次低价平均)": "A04",
    "区间复合一次平均价法": "A08",
    "区间复合平均价浮动法": "A03",
    "简单算术平均值法": "A06",
    "算术平均值下浮法": "A02",
    "区间平均下浮双边曲线算法": "A07",
    "最低价法": "A05",
}

PARAM_LABELS = {
    "w1": "w1",
    "w2": "w2",
    "区间下限": "lower_limit",
    "区间上限": "upper_limit",
    "a": "a",
    "b": "b",
    "n1": "n1",
    "n2": "n2",
    "c": "c",
    "n": "n",
    "m（报价>=基准价）": "m_high",
    "m（报价<基准价）": "m_low",
    "m": "m",
    "F1": "F1",
    "F2": "F2",
    "限价": "ceiling",
    "包限价": "ceiling",
    "价格分占比": "price_weight",
}


def decimal_value(value: object) -> Decimal:
    return Decimal(str(value))


def template_rows(sheet: Worksheet) -> list[tuple[str, str, Decimal]]:
    rows = []
    for row in range(2, sheet.max_row + 1):
        bidder_name = sheet.cell(row, 1).value
        bid_price = sheet.cell(row, 2).value
        score = sheet.cell(row, 3).value
        if bidder_name is None or bid_price is None or score is None:
            continue
        if not str(bidder_name).startswith("投标人"):
            continue
        rows.append((str(bidder_name), str(bid_price), decimal_value(score)))
    return rows


def template_settings(sheet: Worksheet) -> tuple[dict[str, Decimal], Decimal | None, Decimal | None]:
    params: dict[str, Decimal] = {}
    ceiling_price = None
    price_weight = None
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        for index, value in enumerate(values[:-1]):
            if not isinstance(value, str) or value not in PARAM_LABELS:
                continue
            next_value = values[index + 1]
            if next_value is None:
                continue
            mapped = PARAM_LABELS[value]
            if mapped == "ceiling":
                ceiling_price = decimal_value(next_value)
            elif mapped == "price_weight":
                price_weight = decimal_value(next_value)
            else:
                params[mapped] = decimal_value(next_value)
    return params, ceiling_price, price_weight


def template_benchmark(sheet: Worksheet) -> Decimal:
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        for index, value in enumerate(values[:-1]):
            if not isinstance(value, str):
                continue
            if (
                "基准价" in value
                and "折扣率" not in value
                and not value.startswith("m")
                and values[index + 1] is not None
            ):
                return decimal_value(values[index + 1])
    raise AssertionError(f"{sheet.title} missing benchmark value")


def calculate_sheet(sheet_name: str, method_code: str):
    workbook = load_workbook(TEMPLATE, data_only=True, read_only=True)
    sheet = workbook[sheet_name]
    rows = template_rows(sheet)
    params, ceiling_price, price_weight = template_settings(sheet)
    payload = CalculationInput(
        project=ProjectInput(
            tender_no="TEMPLATE",
            tender_name=None,
            section_no="S001",
            package_no="P001",
            province="浙江",
            ceiling_price=ceiling_price,
            price_weight=price_weight,
            target_company=None,
        ),
        method_code=method_code,
        params=params,
        bids=[BidInput(bidder_name=name, bid_price=price, index=index) for index, (name, price, _) in enumerate(rows)],
        source="xlsx_template",
    )
    return sheet, rows, calculate(payload)


@pytest.mark.parametrize("sheet_name,method_code", list(SHEET_METHODS.items()))
def test_v12_template_benchmark_and_scores_match_cached_excel(sheet_name: str, method_code: str) -> None:
    sheet, rows, result = calculate_sheet(sheet_name, method_code)

    assert abs(result.benchmark_price - template_benchmark(sheet)) <= BENCHMARK_TOLERANCE

    result_rows = {row.bidder_name: row for row in result.rows}
    for bidder_name, _, expected_score in rows:
        actual_score = result_rows[bidder_name].score
        assert actual_score is not None
        assert abs(actual_score - expected_score) <= SCORE_TOLERANCE, bidder_name
