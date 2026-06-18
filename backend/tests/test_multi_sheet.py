from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.schemas.calculation import (
    BidPayload,
    CalculateRequest,
    MultiCalculateRequest,
    ProjectPayload,
    SheetCalculatePayload,
)
from app.schemas.mappers import request_to_engine, result_to_response
from app.services.excel import import_bids_from_xlsx_multi, calculate_multi_sheets
from pricing_engine.registry import calculate


def _create_multi_sheet_xlsx() -> bytes:
    wb = Workbook()
    # Sheet1: 浙江 A05
    ws1 = wb.active
    ws1.title = "浙江-包1"
    ws1.append(["投标人名称", "投标价格"])
    ws1.append(["公司A", "100"])
    ws1.append(["公司B", "200"])
    # Sheet2: 江苏 A01
    ws2 = wb.create_sheet("江苏-包2")
    ws2.append(["投标人名称", "投标价格"])
    ws2.append(["公司C", "150"])
    ws2.append(["公司D", "250"])
    ws2.append(["公司E", "不开标"])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_multi_sheet_import() -> None:
    result = import_bids_from_xlsx_multi(_create_multi_sheet_xlsx())
    assert len(result.sheets) == 2
    assert result.sheets[0].name == "浙江-包1"
    assert len(result.sheets[0].rows) == 2
    assert result.sheets[0].rows[0].bidder_name == "公司A"
    assert result.sheets[0].rows[0].bid_price == "100"
    assert result.sheets[1].name == "江苏-包2"
    assert len(result.sheets[1].rows) == 3
    assert result.sheets[1].rows[2].bid_price == "不开标"
    assert result.sheets[1].rows[2].bidder_name == "公司E"


def test_multi_sheet_calculate() -> None:
    payload = MultiCalculateRequest(sheets=[
        SheetCalculatePayload(
            name="浙江-包1",
            project=ProjectPayload(
                tender_no="T001", section_no="S001", package_no="P001",
                province="浙江", ceiling_price="200", price_weight="0.3",
                target_company="公司A",
            ),
            method_code="A05",
            params={},
            bids=[BidPayload(bidder_name="公司A", bid_price="100"),
                  BidPayload(bidder_name="公司B", bid_price="200")],
        ),
        SheetCalculatePayload(
            name="江苏-包2",
            project=ProjectPayload(
                tender_no="T002", section_no="S002", package_no="P002",
                province="江苏", ceiling_price="300", price_weight="0.4",
                target_company="公司C",
            ),
            method_code="A01",
            params={"w1": "0.5", "w2": "0.5", "n1": "1", "c": "0.01", "n2": "1", "round_scale": "1"},
            bids=[BidPayload(bidder_name="公司C", bid_price="150"),
                  BidPayload(bidder_name="公司D", bid_price="250"),
                  BidPayload(bidder_name="公司E", bid_price="不开标")],
        ),
    ])
    result = calculate_multi_sheets(payload)
    assert len(result.results) == 2
    assert result.results[0].name == "浙江-包1"
    assert result.results[0].method_code == "A05"
    assert float(result.results[0].benchmark_price) > 0
    assert result.results[1].name == "江苏-包2"
    assert float(result.results[1].benchmark_price) > 0


def test_multi_sheet_float_direction() -> None:
    """A01 下浮方向：round_scale 应转为 float_direction，和多表保持一致"""
    bids = [BidPayload(bidder_name="公司A", bid_price="180"),
            BidPayload(bidder_name="公司B", bid_price="190")]
    project = ProjectPayload(tender_no="T", section_no="S", package_no="P",
                             province="浙江", ceiling_price="200", price_weight="0.4",
                             target_company="公司A")

    # 单表发法：round_scale=4 + float_direction=-1
    r1 = calculate(request_to_engine(CalculateRequest(
        project=project, method_code="A01",
        params={"w1": "0.5", "w2": "0.5", "n1": "1", "c": "0.01", "n2": "1", "round_scale": "4", "float_direction": "-1"},
        bids=bids, source="test",
    )))

    # 多表发法：round_scale=-1（前端转成 float_direction 后发）
    r2 = calculate(request_to_engine(CalculateRequest(
        project=project, method_code="A01",
        params={"w1": "0.5", "w2": "0.5", "n1": "1", "c": "0.01", "n2": "1", "float_direction": "-1"},
        bids=bids, source="test",
    )))

    sr = result_to_response(r1)
    mr = result_to_response(r2)
    assert sr.benchmark_price == mr.benchmark_price, \
        f"float_direction 不一致: 单表={sr.benchmark_price} 多表={mr.benchmark_price}"


def _make_calc_payload(bids: list[BidPayload], name: str = "test") -> tuple[CalculateRequest, SheetCalculatePayload]:
    """Create identical single-sheet and multi-sheet payloads for comparison."""
    project = ProjectPayload(
        tender_no="T001", section_no="S001", package_no="P001",
        province="浙江", ceiling_price="200", price_weight="0.4",
        target_company=bids[0].bidder_name,
    )
    single = CalculateRequest(
        project=project, method_code="A01",
        params={"w1": "0.5", "w2": "0.5", "n1": "1", "c": "0.01", "n2": "1", "round_scale": "1"},
        bids=bids, source="test",
    )
    multi = SheetCalculatePayload(
        name=name, project=project, method_code="A01",
        params={"w1": "0.5", "w2": "0.5", "n1": "1", "c": "0.01", "n2": "1", "round_scale": "1"},
        bids=bids, source="test",
    )
    return single, multi


def test_single_vs_multi_result_match() -> None:
    """单 sheet 和 多 sheet 对同一份数据测算结果必须一致"""
    bids = [
        BidPayload(bidder_name="公司A", bid_price="180"),
        BidPayload(bidder_name="公司B", bid_price="190"),
        BidPayload(bidder_name="公司C", bid_price="200"),
        BidPayload(bidder_name="公司D", bid_price="210"),
        BidPayload(bidder_name="公司E", bid_price="不开标"),
    ]
    single_req, multi_sheet = _make_calc_payload(bids)
    single_resp = result_to_response(calculate(request_to_engine(single_req)))
    multi_result = calculate_multi_sheets(MultiCalculateRequest(sheets=[multi_sheet]))
    multi_resp = multi_result.results[0]

    assert multi_resp.name == "test"
    assert multi_resp.method_code == single_resp.method_code
    assert multi_resp.method_name == single_resp.method_name
    assert multi_resp.benchmark_price == single_resp.benchmark_price
    assert multi_resp.discount_rate == single_resp.discount_rate
    assert multi_resp.bidder_count == single_resp.bidder_count
    assert multi_resp.effective_count == single_resp.effective_count
    assert multi_resp.target.score == single_resp.target.score
    assert multi_resp.target.rank == single_resp.target.rank
    assert multi_resp.target.score_gap == single_resp.target.score_gap
    for mr, sr in zip(multi_resp.rows, single_resp.rows):
        assert mr.score == sr.score
        assert mr.rank == sr.rank


def test_multi_sheet_import_calculate_roundtrip() -> None:
    xlsx = _create_multi_sheet_xlsx()
    imported = import_bids_from_xlsx_multi(xlsx)
    assert len(imported.sheets) == 2

    payload = MultiCalculateRequest(sheets=[
        SheetCalculatePayload(
            name=s.name,
            project=ProjectPayload(
                tender_no=f"T00{i+1}", section_no=f"S00{i+1}", package_no=f"P00{i+1}",
                province="浙江", ceiling_price="500", price_weight="0.3",
                target_company=s.rows[0].bidder_name,
            ),
            method_code="A05",
            params={},
            bids=s.rows,
        )
        for i, s in enumerate(imported.sheets)
    ])
    result = calculate_multi_sheets(payload)
    assert len(result.results) == 2
    for r in result.results:
        assert float(r.benchmark_price) > 0
        assert len(r.rows) > 0
